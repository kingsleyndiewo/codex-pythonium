from openpyxl import load_workbook
import matplotlib.pyplot as plot
import numpy as np
from datetime import date

# function to fetch workbook data and return sanitized dictionary
def getExcelData(filePath = 'Sample Data.xlsx'):
	# load the file
	revenue = load_workbook(filename = filePath, read_only=True)
	# we only need the first worksheet
	workSheet = revenue.get_sheet_by_name(revenue.sheetnames[0])
	data = {'revenue': {}, 'regions': []}
	products = None
	# loop through rows
	for r in workSheet.iter_rows():
		# r is a tuple of ReadOnlyCell objects
		# (Region, Detergent, Cooking Oil, Sugar, Cereal, Total)
		regionName = str(r[0].value)
		# Skip header row
		if regionName == 'REGION':
			# get the product names
			products = r[1].value, r[2].value, r[3].value, r[4].value
			continue
		# add region name to regions
		data['regions'].append(regionName)
		# add product to dictionary
		for i, p in enumerate(products):
			if p not in data['revenue']:
				data['revenue'][p] = []
			# add the product revenue of current region to the list for this product
			data['revenue'][p].append(r[i + 1].value)
	# we should have all the revenue data now
	return data

# function to plot a compound bar graph of all regions on x and product revenues on Y for the totals
def plotRevCBG(revenueDB):
	# get the region names
	regions = revenueDB['regions']
	# get the product names
	products = list(revenueDB['revenue'].keys())
	# sort the names
	products.sort()
	# number of x elements
	N = len(regions)
	# get the locations for the x-axis
	locs = np.arange(N)
	# width of bars
	width = 0.35
	largest = 0
	# create a plot for each product
	subPlots = []
	valueLists = []
	for row, product in enumerate(products):
		# get the value list
		valuesList = tuple(revenueDB['revenue'][product])
		valueLists.append(valuesList)
		if row == 0:
			# this is the first one
			subPlot = plot.bar(locs, valuesList, width)
		else:
			# put the previous data below it
			subPlot = plot.bar(locs, valuesList, width, bottom=valueLists[row - 1])
		# add to the list
		subPlots.append(subPlot)
	# get maximum value for Y-axis (the tallest bar)
	sums = [sum(x) for x in zip(*valueLists)]
	largest = max(sums)
	# add 10% on top
	largest *= 1.1
	# plot basics
	plot.ylabel('Sales Revenue (KES)')
	plot.title('Monthly sales revenue for Foo Traders by Region and Product')
	xticks = regions
	plot.xticks(locs, xticks)
	# make ticks from 0 to the largest value possible when stacked
	plot.yticks(np.arange(0, largest, (largest - (largest % 10)) / 10))
	plot.legend(tuple([p[0] for p in subPlots]), tuple([n for n in products]))
	# show the plot
	plot.show()

if __name__ == '__main__':
	# get the file path from user
	#filePath = raw_input("Please enter the path to the Excel file: ")
	# get the data
	revenueData = getExcelData()
	# plot a compound bar graph
	plotRevCBG(revenueData)
	